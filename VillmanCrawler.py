import re
import requests
import urllib.parse
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from lxml import etree

def get_html_from(url):
	r = requests.get(url, verify=False)
	r.encoding = 'cp950'
	return etree.HTML(r.text)

def fill_color(row, column, color):
	my_color = openpyxl.styles.colors.Color(rgb=color)
	my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_color)
	ws.cell(row=row, column=column).fill = my_fill	

def fill_none(row, column):
	my_fill = openpyxl.styles.fills.PatternFill(fill_type=None)
	ws.cell(row=row, column=column).fill = my_fill	

WORK_BOOK_NAME = "Villman.xlsx"

wb = load_workbook(WORK_BOOK_NAME)

url = "https://villman.com"
url_page = url + "/Category/Notebook-PCs/%s"
ws = wb.worksheets[0]
dws = wb.worksheets[1]

bottom_row = 1
discard_bottom_row = 1
exist_data = {}
active_data = {}

while (True):
	bottom_row += 1
	fill_none(row=bottom_row, column=1)
	name = ws.cell(row=bottom_row, column=1).value
	if name == None:
		break;
	else:
		exist_data[name] = bottom_row

while (True):
	discard_bottom_row += 1
	name = dws.cell(row=discard_bottom_row, column=1).value
	if name == None:
		break;

page = 0

while page < 500:

	r = requests.get(url_page % page, verify=False)
	html = etree.HTML(r.text)

	prod_link_list = html.xpath("*//a[@class='prod_link']")

	for a in prod_link_list:
		name = a.text
		if name in exist_data:
			row = exist_data[name]
			active_data[name] = row
			del exist_data[name]
		else: 
			row = bottom_row
			bottom_row += 1
			ws.cell(row=row, column=1).value = name
			fill_color(row=row, column=1, color='00FFFF00')
		prod_link = a.get("href")
		ws.cell(row=row, column=2).value = '=HYPERLINK("{}", "{}")'.format(url + prod_link, "Link")

	page += 50

for key in exist_data.keys():
	row = exist_data[key]
	dws.cell(row=discard_bottom_row, column=1).value = ws.cell(row=row, column=1).value
	for col in range(3, 10):
		dws.cell(row=discard_bottom_row, column=col - 1).value = ws.cell(row=row, column=col).value
	ws.delete_rows(row)
	bottom_row -= 1
	discard_bottom_row += 1

wb.save(WORK_BOOK_NAME)
